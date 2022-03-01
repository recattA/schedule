import math

from openpyxl import Workbook,  load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Fill, PatternFill, NamedStyle, Side, Border, Alignment

data = {
    "Кучеренко": {
        "Понедельник": " ",
        "Вторник": " ",
        "Среда": "В 1402-3",
        "Четверг": " ",
        "Пятница": "В 1402-3, В 1402-1",
        "Суббота": "А 514а"
    },
    "Шабалина О.А.": {
        "Понедельник": "В 1402-1",
        "Вторник": "В 1402-1",
        "Среда": "А-514а,б",
        "Четверг": "В-1402-3",
        "Пятница": "В 1402-3, В-1405, В 1005",
        "Суббота": " "
    },
    "Щербаков М.В.": {
        "Понедельник": "Дистанционно",
        "Вторник": "В-1301",
        "Среда": "В-1302, В-1401",
        "Четверг": " ",
        "Пятница": "В-1302а, В-1302",
        "Суббота": "Дистанционно"
    },
    "Драгунов С.Е.": {
        "Понедельник": "В 1402-2",
        "Вторник": "В-1402-2, В-1401, В-403",
        "Среда": "В 1402-2, В 403",
        "Четверг": " ",
        "Пятница": " ",
        "Суббота": " "
    },
    "Аль-Гунаид Мохаммед": {
        "Понедельник": " ",
        "Вторник": " ",
        "Среда": "В 1405",
        "Четверг": "В 1401",
        "Пятница": " ",
        "Суббота": " "
    },
    "Соколов А.А.": {
        "Понедельник": "В-1402-3, В-801, Дистанционно",
        "Вторник": "В-1402-3",
        "Среда": "В-1402-3, В-1003",
        "Четверг": "409",
        "Пятница": "Дистанционно, В-1402-3, В-1003",
        "Суббота": "В-1402-3, А 600а"
    },
    "Фоменков С.А.": {
        "Понедельник": " ",
        "Вторник": "Дистанционно",
        "Среда": "В-502, В-801",
        "Четверг": "В-801, Дистанционно",
        "Пятница": "Дистанционно",
        "Суббота": " "
    },
    "Козина С.А.": {
        "Понедельник": " ",
        "Вторник": "В-1402-1",
        "Среда": "В-1402-1",
        "Четверг": " ",
        "Пятница": "В-1402-1",
        "Суббота": "В-1402-1"
    },
    "Катаев А.В.": {
        "Понедельник": "В-1402-1",
        "Вторник": "В 1402-3, Дистанционно",
        "Среда": "В 1402-1, В 501",
        "Четверг": "В 1402-3",
        "Пятница": " ",
        "Суббота": " "
    },
    "Матохина А.В.": {
        "Понедельник": "В-1402-3, В 1401",
        "Вторник": "В 1402-3, В 1401, В 403",
        "Среда": "Дистанционно, В 1401",
        "Четверг": "В 1402-1",
        "Пятница": "В 1401, В 1005, Дистанционно",
        "Суббота": "В-1402-3"
    },
    "Яновский Т.А.": {
        "Понедельник": " ",
        "Вторник": "Дистанционно",
        "Среда": " ",
        "Четверг": "В-1402-1",
        "Пятница": " ",
        "Суббота": " "
    },
    "Садовникова Н.П.": {
        "Понедельник": " ",
        "Вторник": "Дистанционно",
        "Среда": " ",
        "Четверг": "В 1402-1, В-1402-3",
        "Пятница": "В 1005, В-1401",
        "Суббота": " "
    },
    "Иванченко С.": {
        "Понедельник": "В-1402-3",
        "Вторник": " ",
        "Среда": " ",
        "Четверг": "В-1402-3",
        "Пятница": " ",
        "Суббота": " "
    },
    "Кизим А.В.": {
        "Понедельник": " ",
        "Вторник": " ",
        "Среда": "В-1402-1",
        "Четверг": " ",
        "Пятница": "В-1402-3",
        "Суббота": " "
    },
    "Черемисинов С.В.": {
        "Понедельник": " ",
        "Вторник": " ",
        "Среда": " ",
        "Четверг": " ",
        "Пятница": "В-1405",
        "Суббота": " "
    },
    "Коробкин Д.М.": {
        "Понедельник": " ",
        "Вторник": " ",
        "Среда": " ",
        "Четверг": " ",
        "Пятница": " ",
        "Суббота": "Дистанционно"
    },
}

wb = Workbook()
ws = wb.active

headings = ['Фамилия/День недели'] + list(data['Кучеренко'].keys())
ws.append(headings)

for person in data:
    grades = list(data[person].values())
    ws.append([person] + grades)



for i, rowOfCellObjects in enumerate(ws['A1':'G17']):
    for n, cellObj in enumerate(rowOfCellObjects):
        cellObj.font = Font(size=20)


col_width = []
for i in range(len(next(ws.iter_rows()))):
    col_letter = get_column_letter(i + 1)

    minimum_width = 50
    current_width = ws.column_dimensions[col_letter].width
    if not current_width or current_width < minimum_width:
        ws.column_dimensions[col_letter].width = minimum_width

    col_width.append(ws.column_dimensions[col_letter].width)

for i, row in enumerate(ws):
    default_height = 26

    multiples_of_font_size = [default_height]
    for j, cell in enumerate(row):
        vertical = "center"
        horizontal = "center"
        if cell.value is not None:
            mul = 0
            for v in str(cell.value).split('\n'):
                mul += math.ceil(len(v) / col_width[j]) * cell.font.size

            if mul > 0:
                multiples_of_font_size.append(mul)

        cell.alignment = Alignment(vertical=vertical, horizontal=horizontal)

    original_height = ws.row_dimensions[i + 1].height
    if original_height is None:
        original_height = default_height

    new_height = max(multiples_of_font_size)
    if original_height < new_height:
        ws.row_dimensions[i + 1].height = new_height



wb.save("ТестОтчёт.xlsx")
wb.close()