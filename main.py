import math

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Fill, PatternFill, NamedStyle, Side, Border, Alignment
from math import ceil



wb = Workbook()
# wb = openpyxl.open("ОН_Магистратура_2 курс ФЭВТ.xlsx")
ws = wb.active

# Стилизовать шрифты
ws['A1'] = "Расписание занятий преподавателей"
ws.merge_cells("A1:BA3")
ws['A1'].font = Font(size=36)
ws['A1'].alignment = Alignment(horizontal='center')

ws['A4'] = "Ф.И.О."
ws.merge_cells("A4:D6")
ws['A4'].font = Font(size=36)
ws['A4'].alignment = Alignment(horizontal='center')
ws['E4'] = "Неделя Часы"
ws.merge_cells("E4:E6")




# дни недели
ws['F4'] = "Понедельник"
ws.merge_cells("F4:M4")
ws['F4'].alignment = Alignment(horizontal='center')
ws['N4'] = "Вторник"
ws.merge_cells("N4:U4")
ws['N4'].alignment = Alignment(horizontal='center')
ws['V4'] = "Среда"
ws.merge_cells("V4:AC4")
ws['V4'].alignment = Alignment(horizontal='center')
ws['AD4'] = "Четверг"
ws.merge_cells("AD4:AK4")
ws['AD4'].alignment = Alignment(horizontal='center')
ws['AL4'] = "Пятница"
ws.merge_cells("AL4:AS4")
ws['AL4'].alignment = Alignment(horizontal='center')
ws['AT4'] = "Суббота"
ws.merge_cells("AT4:BA4")
ws['AT4'].alignment = Alignment(horizontal='center')

# преподаватели
ws['A7'] = "Кучеренко"
ws.merge_cells("A7:D10")
ws['A7'].font = Font(size=48)
ws['A7'].alignment = Alignment(horizontal='center')
ws['A11'] = "Шабалина О.А."
ws.merge_cells("A11:D14")
ws['A11'].font = Font(size=48)
ws['A11'].alignment = Alignment(horizontal='center')
ws['A15'] = "Щербаков М.В."
ws.merge_cells("A15:D18")
ws['A15'].font = Font(size=48)
ws['A15'].alignment = Alignment(horizontal='center')
ws['A19'] = "Драгунов С.Е."
ws.merge_cells("A19:D22")
ws['A19'].font = Font(size=48)
ws['A19'].alignment = Alignment(horizontal='center')
ws['A23'] = "Аль-Гунаид Мохаммед"
ws.merge_cells("A23:D26")
ws['A23'].font = Font(size=40)
ws['A23'].alignment = Alignment(horizontal='center')
ws['A27'] = "Соколов А.А."
ws.merge_cells("A27:D30")
ws['A27'].font = Font(size=48)
ws['A27'].alignment = Alignment(horizontal='center')
ws['A31'] = "Фоменков С.А."
ws.merge_cells("A31:D34")
ws['A31'].font = Font(size=48)
ws['A31'].alignment = Alignment(horizontal='center')
ws['A35'] = "Козина С.А."
ws.merge_cells("A35:D38")
ws['A35'].font = Font(size=48)
ws['A35'].alignment = Alignment(horizontal='center')
ws['A39'] = "Катаев А.В."
ws.merge_cells("A39:D42")
ws['A39'].font = Font(size=48)
ws['A39'].alignment = Alignment(horizontal='center')
ws['A43'] = "Матохина А.В."
ws.merge_cells("A43:D46")
ws['A43'].font = Font(size=48)
ws['A43'].alignment = Alignment(horizontal='center')
ws['A47'] = "Яновский Т.А."
ws.merge_cells("A47:D50")
ws['A47'].font = Font(size=48)
ws['A47'].alignment = Alignment(horizontal='center')
ws['A51'] = "Садовникова Н.П."
ws.merge_cells("A51:D54")
ws['A51'].font = Font(size=48)
ws['A51'].alignment = Alignment(horizontal='center')
ws['A55'] = "Иванченко С."
ws.merge_cells("A55:D58")
ws['A55'].font = Font(size=48)
ws['A55'].alignment = Alignment(horizontal='center')
ws['A59'] = "Кизим А.В."
ws.merge_cells("A59:D62")
ws['A59'].font = Font(size=48)
ws['A59'].alignment = Alignment(horizontal='center')
ws['A63'] = "Черемисинов С.В."
ws.merge_cells("A63:D66")
ws['A63'].font = Font(size=48)
ws['A63'].alignment = Alignment(horizontal='center')
ws['A67'] = "Коробкин Д.М."
ws.merge_cells("A67:D70")
ws['A67'].font = Font(size=48)
ws['A67'].alignment = Alignment(horizontal='center')


# недели
ws['E7'] = "1"
ws.merge_cells("E7:E8")
ws['E7'].alignment = Alignment(horizontal='center')
ws['E7'].font = Font(size=22)
ws['E9'] = "2"
ws.merge_cells("E9:E10")
ws['E9'].alignment = Alignment(horizontal='center')
ws['E9'].font = Font(size=22)
ws['E11'] = "1"
ws.merge_cells("E11:E12")
ws['E11'].alignment = Alignment(horizontal='center')
ws['E11'].font = Font(size=22)
ws['E13'] = "2"
ws.merge_cells("E13:E14")
ws['E13'].alignment = Alignment(horizontal='center')
ws['E13'].font = Font(size=22)
ws['E15'] = "1"
ws.merge_cells("E15:E16")
ws['E15'].alignment = Alignment(horizontal='center')
ws['E15'].font = Font(size=22)
ws['E17'] = "2"
ws.merge_cells("E17:E18")
ws['E17'].alignment = Alignment(horizontal='center')
ws['E17'].font = Font(size=22)
ws['E19'] = "1"
ws.merge_cells("E19:E20")
ws['E19'].alignment = Alignment(horizontal='center')
ws['E19'].font = Font(size=22)
ws['E21'] = "2"
ws.merge_cells("E21:E22")
ws['E21'].alignment = Alignment(horizontal='center')
ws['E21'].font = Font(size=22)
ws['E23'] = "1"
ws.merge_cells("E23:E24")
ws['E23'].alignment = Alignment(horizontal='center')
ws['E23'].font = Font(size=22)
ws['E25'] = "2"
ws.merge_cells("E25:E26")
ws['E25'].alignment = Alignment(horizontal='center')
ws['E25'].font = Font(size=22)
ws['E27'] = "1"
ws.merge_cells("E27:E28")
ws['E27'].alignment = Alignment(horizontal='center')
ws['E27'].font = Font(size=22)
ws['E29'] = "2"
ws.merge_cells("E29:E30")
ws['E29'].alignment = Alignment(horizontal='center')
ws['E29'].font = Font(size=22)
ws['E31'] = "1"
ws.merge_cells("E31:E32")
ws['E31'].alignment = Alignment(horizontal='center')
ws['E31'].font = Font(size=22)
ws['E33'] = "2"
ws.merge_cells("E33:E34")
ws['E33'].alignment = Alignment(horizontal='center')
ws['E33'].font = Font(size=22)
ws['E35'] = "1"
ws.merge_cells("E35:E36")
ws['E35'].alignment = Alignment(horizontal='center')
ws['E35'].font = Font(size=22)
ws['E37'] = "2"
ws.merge_cells("E37:E38")
ws['E37'].alignment = Alignment(horizontal='center')
ws['E37'].font = Font(size=22)
ws['E39'] = "1"
ws.merge_cells("E39:E40")
ws['E39'].alignment = Alignment(horizontal='center')
ws['E39'].font = Font(size=22)
ws['E41'] = "2"
ws.merge_cells("E41:E42")
ws['E41'].alignment = Alignment(horizontal='center')
ws['E41'].font = Font(size=22)
ws['E43'] = "1"
ws.merge_cells("E43:E44")
ws['E43'].alignment = Alignment(horizontal='center')
ws['E43'].font = Font(size=22)
ws['E45'] = "2"
ws.merge_cells("E45:E46")
ws['E45'].alignment = Alignment(horizontal='center')
ws['E45'].font = Font(size=22)
ws['E47'] = "1" #
ws.merge_cells("E47:E48")
ws['E47'].alignment = Alignment(horizontal='center')
ws['E47'].font = Font(size=22)
ws['E49'] = "2"
ws.merge_cells("E49:E50")
ws['E49'].alignment = Alignment(horizontal='center')
ws['E49'].font = Font(size=22)
ws['E51'] = "1"
ws.merge_cells("E51:E52")
ws['E51'].alignment = Alignment(horizontal='center')
ws['E51'].font = Font(size=22)
ws['E53'] = "2"
ws.merge_cells("E53:E54")
ws['E53'].alignment = Alignment(horizontal='center')
ws['E53'].font = Font(size=22)
ws['E55'] = "1"
ws.merge_cells("E55:E56")
ws['E55'].alignment = Alignment(horizontal='center')
ws['E55'].font = Font(size=22)
ws['E57'] = "2"
ws.merge_cells("E57:E58")
ws['E57'].alignment = Alignment(horizontal='center')
ws['E57'].font = Font(size=22)
ws['E59'] = "1"
ws.merge_cells("E59:E60")
ws['E59'].alignment = Alignment(horizontal='center')
ws['E59'].font = Font(size=22)
ws['E61'] = "2"
ws.merge_cells("E61:E62")
ws['E61'].alignment = Alignment(horizontal='center')
ws['E61'].font = Font(size=22)
ws['E63'] = "1"
ws.merge_cells("E63:E64")
ws['E63'].alignment = Alignment(horizontal='center')
ws['E63'].font = Font(size=22)
ws['E65'] = "2"
ws.merge_cells("E65:E66")
ws['E65'].alignment = Alignment(horizontal='center')
ws['E65'].font = Font(size=22)
ws['E67'] = "1"
ws.merge_cells("E67:E68")
ws['E67'].alignment = Alignment(horizontal='center')
ws['E67'].font = Font(size=22)
ws['E69'] = "2"
ws.merge_cells("E69:E70")
ws['E69'].alignment = Alignment(horizontal='center')
ws['E69'].font = Font(size=22)

# часы
# понедельник
ws['F5'] = "1-2"
ws.merge_cells("F5:F6")
ws['F5'].font = Font(size=20)
ws['G5'] = "3-4"
ws.merge_cells("G5:G6")
ws['G5'].font = Font(size=20)
ws['H5'] = "5-6"
ws.merge_cells("H5:H6")
ws['H5'].font = Font(size=20)
ws['I5'] = "7-8"
ws.merge_cells("I5:I6")
ws['I5'].font = Font(size=20)
ws['J5'] = "9-10"
ws.merge_cells("J5:J6")
ws['J5'].font = Font(size=20)
ws['K5'] = "11-12"
ws.merge_cells("K5:K6")
ws['K5'].font = Font(size=20)
ws['L5'] = "13-14"
ws.merge_cells("L5:L6")
ws['L5'].font = Font(size=20)
ws['M5'] = "15-16"
ws.merge_cells("M5:M6")
ws['M5'].font = Font(size=20)

# вторник
ws['N5'] = "1-2"
ws.merge_cells("N5:N6")
ws['N5'].font = Font(size=20)
ws['O5'] = "3-4"
ws.merge_cells("O5:O6")
ws['O5'].font = Font(size=20)
ws['P5'] = "5-6"
ws.merge_cells("P5:P6")
ws['P5'].font = Font(size=20)
ws['Q5'] = "7-8"
ws.merge_cells("Q5:Q6")
ws['Q5'].font = Font(size=20)
ws['R5'] = "9-10"
ws.merge_cells("R5:R6")
ws['R5'].font = Font(size=20)
ws['S5'] = "11-12"
ws.merge_cells("S5:S6")
ws['S5'].font = Font(size=20)
ws['T5'] = "13-14"
ws.merge_cells("T5:T6")
ws['T5'].font = Font(size=20)
ws['U5'] = "15-16"
ws.merge_cells("U5:U6")
ws['U5'].font = Font(size=20)

# среда
ws['V5'] = "1-2"
ws.merge_cells("V5:V6")
ws['V5'].font = Font(size=20)
ws['W5'] = "3-4"
ws.merge_cells("W5:W6")
ws['W5'].font = Font(size=20)
ws['X5'] = "5-6"
ws.merge_cells("X5:X6")
ws['X5'].font = Font(size=20)
ws['Y5'] = "7-8"
ws.merge_cells("Y5:Y6")
ws['Y5'].font = Font(size=20)
ws['Z5'] = "9-10"
ws.merge_cells("Z5:Z6")
ws['Z5'].font = Font(size=20)
ws['AA5'] = "11-12"
ws.merge_cells("AA5:AA6")
ws['AA5'].font = Font(size=20)
ws['AB5'] = "13-14"
ws.merge_cells("AB5:AB6")
ws['AB5'].font = Font(size=20)
ws['AC5'] = "15-16"
ws.merge_cells("AC5:AC6")
ws['AC5'].font = Font(size=20)

# четверг
ws['AD5'] = "1-2"
ws.merge_cells("AD5:AD6")
ws['AD5'].font = Font(size=20)
ws['AE5'] = "3-4"
ws.merge_cells("AE5:AE6")
ws['AE5'].font = Font(size=20)
ws['AF5'] = "5-6"
ws.merge_cells("AF5:AF6")
ws['AF5'].font = Font(size=20)
ws['AG5'] = "7-8"
ws.merge_cells("AG5:AG6")
ws['AG5'].font = Font(size=20)
ws['AH5'] = "9-10"
ws.merge_cells("AH5:AH6")
ws['AH5'].font = Font(size=20)
ws['AI5'] = "11-12"
ws.merge_cells("AI5:AI6")
ws['AI5'].font = Font(size=20)
ws['AJ5'] = "13-14"
ws.merge_cells("AJ5:AJ6")
ws['AJ5'].font = Font(size=20)
ws['AK5'] = "15-16"
ws.merge_cells("AK5:AK6")
ws['AK5'].font = Font(size=20)

# пятница
ws['AL5'] = "1-2"
ws.merge_cells("AL5:AL6")
ws['AL5'].font = Font(size=20)
ws['AM5'] = "3-4"
ws.merge_cells("AM5:AM6")
ws['AM5'].font = Font(size=20)
ws['AN5'] = "5-6"
ws.merge_cells("AN5:AN6")
ws['AN5'].font = Font(size=20)
ws['AO5'] = "7-8"
ws.merge_cells("AO5:AO6")
ws['AO5'].font = Font(size=20)
ws['AP5'] = "9-10"
ws.merge_cells("AP5:AP6")
ws['AP5'].font = Font(size=20)
ws['AQ5'] = "11-12"
ws.merge_cells("AQ5:AQ6")
ws['AQ5'].font = Font(size=20)
ws['AR5'] = "13-14"
ws.merge_cells("AR5:AR6")
ws['AR5'].font = Font(size=20)
ws['AS5'] = "15-16"
ws.merge_cells("AS5:AS6")
ws['AS5'].font = Font(size=20)

# суббота
ws['AT5'] = "1-2"
ws.merge_cells("AT5:AT6")
ws['AT5'].font = Font(size=20)
ws['AU5'] = "3-4"
ws.merge_cells("AU5:AU6")
ws['AU5'].font = Font(size=20)
ws['AV5'] = "5-6"
ws.merge_cells("AV5:AV6")
ws['AV5'].font = Font(size=20)
ws['AW5'] = "7-8"
ws.merge_cells("AW5:AW6")
ws['AW5'].font = Font(size=20)
ws['AX5'] = "9-10"
ws.merge_cells("AX5:AX6")
ws['AX5'].font = Font(size=20)
ws['AY5'] = "11-12"
ws.merge_cells("AY5:AY6")
ws['AY5'].font = Font(size=20)
ws['AZ5'] = "13-14"
ws.merge_cells("AZ5:AZ6")
ws['AZ5'].font = Font(size=20)
ws['BA5'] = "15-16"
ws.merge_cells("BA5:BA6")
ws['BA5'].font = Font(size=20)

#
ws['Z7'] = "В 1402-3"
ws['AA7'] = "В-1402-3"
ws['AT7'] = "А 514а"
ws['AU7'] = "А 514а"
ws['AL9'] = "В 1402-3"
ws['AM9'] = "В 1402-3"
ws['AP9'] = "В 1402-1"
ws['AQ9'] = "В 1402-1"  #
ws['AQ11'] = "В-1405"
ws['AR11'] = "В-1405"
ws['R13'] = "В-1402-1"
ws['S13'] = "В-1402-1"
ws['Z13'] = "А-514а,б"
ws['AA13'] = "А-514а,б"
ws['Q15'] = "В-1301"
ws['R15'] = "В-1301"
ws['AP15'] = "В-1302а"
ws['Y17'] = "В-1302"
ws['AP17'] = "В-1302"
ws['AQ17'] = "В-1302"
ws['R21'] = "В-403"
ws['S21'] = "В-403"
ws['Y21'] = "В 403"
ws['Z25'] = "В-1405"
ws['AA25'] = "В-1405"
ws['I15'] = "Дистанционно"
ws['H17'] = "Дистанционно"
ws['F27'] = "В-1402-3"
ws['G27'] = "В-1402-3"
ws['H27'] = "В-801"
ws['V27'] = "В-1402-3"
ws['W27'] = "В-1402-3"
ws['AF27'] = "409"
ws['AG27'] = "409"
ws['AN27'] = "Дистанционно"
ws['AT27'] = "В-1402-3"
ws['AU27'] = "В-1402-3"
ws['AV27'] = "В-1402-3"
ws['AW27'] = "В-1402-3"
ws['G29'] = "Дистанционно"
ws['N29'] = "В-1402-3"
ws['O29'] = "В-1402-3"
ws['P29'] = "В-1402-3"
ws['Q29'] = "В-1402-3"
ws['V29'] = "В-1402-3"
ws['W29'] = "В-1402-3"
ws['X29'] = "В-1003"
ws['Y29'] = "В-1003"
ws['AL29'] = "Дистанционно"
ws['AN29'] = "В-1402-3"
ws['AO29'] = "В-1402-3"
ws['AP29'] = "В-1003"
ws['AT29'] = "А 600а"
ws['AU29'] = "А 600а"
ws['AV29'] = "А 600а"
ws['AW29'] = "А 600а" #
ws['O31'] = "Дистанционно"
ws['AG31'] = "В-801"
ws['AF31'] = "В-801"
ws['AM31'] = "Дистанционно"
ws['W33'] = "В-502"
ws['X33'] = "В-801"
ws['AE33'] = "Дистанционно"
ws['V35'] = "В-1402-1"
ws['W35'] = "В-1402-1"
ws['X35'] = "В-1402-1"
ws['Y35'] = "В-1402-1"
ws['AT35'] = "В-1402-1"
ws['AU35'] = "В-1402-1"
ws['I29'] = "В-1402-3"
ws['J29'] = "В-1402-3"
ws['N37'] = "В-1402-1"
ws['O37'] = "В-1402-1"
ws['P37'] = "В-1402-1"
ws['Q37'] = "В-1402-1"
ws['V37'] = "В-1402-1"
ws['W37'] = "В-1402-1"
ws['AN37'] = "В-1402-1"
ws['AO37'] = "В-1402-1"
ws['AT37'] = "В-1402-1"
ws['AU37'] = "В-1402-1"
ws['F39'] = "В 1402-1"
ws['G39'] = "В 1402-1"
ws['N39'] = "В 1402-3"
ws['O39'] = "В 1402-3"
ws['AF39'] = "В 1402-3"
ws['AG39'] = "В 1402-3"
ws['AH39'] = "В 1402-3"
ws['AI39'] = "В 1402-3"
ws['F41'] = "В 1401"
ws['G41'] = "В 1401"
ws['P41'] = "Дистанционно"
ws['Y41'] = "В 501"
ws['AD41'] = "В 1402-3"
ws['AE41'] = "В 1402-3"
ws['I43'] = "В-1402-3"
ws['J43'] = "В 1401"
ws['P43'] = "В 1402-3"
ws['Q43'] = "В 1402-3"
ws['Y43'] = "Дистанционно"
ws['AD43'] = "В 1402-1"
ws['AE43'] = "В 1402-1"
ws['AL43'] = "В 1401"
ws['AM43'] = "В 1401"
ws['F45'] = "В 1402-3"
ws['G45'] = "В 1402-3"
ws['H45'] = "В 1402-3"
ws['O45'] = "В 1401"
ws['V45'] = "В 1401"
ws['W45'] = "В 1401"
ws['X45'] = "В 1401"
ws['Y45'] = "В 1401"
ws['Z45'] = "В 1401"
ws['AA45'] = "В 1401"
ws['AM45'] = "Дистанционно"
ws['AT45'] = "В-1402-3"
ws['AU45'] = "В-1402-3"
ws['O11'] = "В-1401"
ws['P11'] = "В-1401"
ws['AM11'] = "В 1402-3"
ws['AF13'] = "В-1402-3"
ws['AG13'] = "В-1402-3"
ws['Q49'] = "Дистанционно"
ws['AD49'] = "В-1402-1"
ws['AE49'] = "В-1402-1"
ws['AF49'] = "В-1402-1"
ws['AG49'] = "В-1402-1"
ws['H11'] = "В 1402-1"
ws['I11'] = "В 1402-1"
ws['J11'] = "В 1402-1"
ws['K11'] = "В 1402-1"
ws['AP13'] = "В 1005"
ws['H39'] = "В 1401"
ws['I39'] = "В 1401"
ws['Q39'] = "Дистанционно"
ws['X41'] = "В 1402-1"
ws['Z41'] = "В 1402-1"
ws['AA41'] = "В 1402-1"
ws['X23'] = "В 1405"
ws['Y23'] = "В 1405"
ws['AF25'] = "В 1401"
ws['AG25'] = "В 1401"
ws['AH25'] = "В 1401"
ws['AI25'] = "В 1401"
ws['X19'] = "В 1402-2"
ws['Y19'] = "В 1402-2"
ws['P21'] = "В 1402-2"
ws['I21'] = "В 1402-2"
ws['V21'] = "В 1402-2"
ws['W21'] = "В 1402-2"
ws['AF51'] = "В 1402-1"
ws['AG51'] = "В 1402-1"
ws['AN51'] = "В 1005"
ws['AN43'] = "В 1005"
ws['AO43'] = "В 1005"
ws['AP43'] = "В 1005"
ws['P45'] = "В 403"
ws['Q45'] = "В 403"
ws['J55'] = "В-1402-3"
ws['K55'] = "В-1402-3"
ws['AD55'] = "В-1402-3"
ws['AE55'] = "В-1402-3"
ws['O19'] = "В-1402-2"
ws['P19'] = "В-1402-2"
ws['Q19'] = "В-1401"
ws['R19'] = "В-1402-2"
ws['S19'] = "В-1402-2"
ws['R43'] = "В-1402-3"
ws['S43'] = "В-1401"
ws['Z59'] = "В-1402-1"
ws['AA59'] = "В-1402-1"
ws['AP61'] = "В-1402-3"
ws['AQ61'] = "В-1402-3"
ws['Z15'] = "В-1401"
ws['AA15'] = "В-1401"
ws['AQ15'] = "В-1402-3"
ws['AU15'] = "Дистанционно"
ws['AP51'] = "В-1401"
ws['AQ51'] = "В-1401"
ws['S53'] = "Дистанционно"
ws['AH53'] = "В-1402-3"
ws['AI53'] = "В-1402-3"
ws['AM65'] = "В-1405"
ws['AU67'] = "Дистанционно"
# черновик



for i, rowOfCellObjects in enumerate(ws['F7':'BA70']):
    for n, cellObj in enumerate(rowOfCellObjects):
        cellObj.font = Font(size=26)




col_width = []
for i in range(len(next(ws.iter_rows()))):
    col_letter = get_column_letter(i + 1)

    minimum_width = 20
    current_width = ws.column_dimensions[col_letter].width
    if not current_width or current_width < minimum_width:
        ws.column_dimensions[col_letter].width = minimum_width

    col_width.append(ws.column_dimensions[col_letter].width)

for i, row in enumerate(ws):
    default_height = 26

    multiples_of_font_size = [default_height]
    for j, cell in enumerate(row):
        vertical = "center"
        horizontal = "center"
        if cell.value is not None:
            mul = 0
            for v in str(cell.value).split('\n'):
                mul += math.ceil(len(v) / col_width[j]) * cell.font.size

            if mul > 0:
                multiples_of_font_size.append(mul)

        cell.alignment = Alignment(vertical=vertical, horizontal=horizontal)

    original_height = ws.row_dimensions[i + 1].height
    if original_height is None:
        original_height = default_height

    new_height = max(multiples_of_font_size)
    if original_height < new_height:
        ws.row_dimensions[i + 1].height = new_height







wb.save('РасписаниеПР.xlsx')
wb.close()
