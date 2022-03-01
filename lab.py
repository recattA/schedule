import math

from openpyxl import Workbook,  load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Fill, PatternFill, NamedStyle, Side, Border, Alignment

data = {
    "1401": {
        "Понедельник": "Матохина А.В.",
        "Вторник": "Драгунов С.Е., Матохина А.В.",
        "Среда": "Щербаков М.В., Матохина А.В.",
        "Четверг": "Аль-Гунаид Мохаммед",
        "Пятница": "Матохина А.В., Садовникова Н.П.",
        "Суббота": " "
    },
    "В 1402-1": {
        "Понедельник": "Катаев А.В., Шабалина О.А.",
        "Вторник": "Козина С.А., Шабалина О.А.",
        "Среда": "Козина С.А., Катаев А.В., Кизим А.В.",
        "Четверг": "Матохина А.В., Яновский Т.А., Садовникова Н.П.",
        "Пятница": "Козина С.А., Кучеренко",
        "Суббота": "Козина С.А."
    },
    "В-1402-2": {
        "Понедельник": "Драгунов С.Е.",
        "Вторник": "Драгунов С.Е.",
        "Среда": "Драгунов С.Е.",
        "Четверг": " ",
        "Пятница": " ",
        "Суббота": " "
    },
    "В-1402-3": {
        "Понедельник": "Соколов А.А., Матохина А.В., Иванченко С.",
        "Вторник": "Соколов А.А., Катаев А.В., Матохина А.В.",
        "Среда": "Соколов А.А., Кучеренко",
        "Четверг": "Шабалина О.А., Садовникова Н.П., Иванченко С., Катаев А.В.",
        "Пятница": "Соколов А.А., Кизим А.В., Кучеренко, Шабалина О.А.",
        "Суббота": "Соколов А.А., Матохина А.В."
    },
    "В-1405": {
        "Понедельник": " ",
        "Вторник": " ",
        "Среда": "Аль-Гунаид Мохаммед",
        "Четверг": " ",
        "Пятница": "Шабалина О.А., Черемисинов С.В.",
        "Суббота": "А 514а"
    },
}


wb = Workbook()
ws = wb.active


headings = ['Аудитория/День недели'] + list(data['1401'].keys())
ws.append(headings)

for person in data:
    grades = list(data[person].values())
    ws.append([person] + grades)

for i, rowOfCellObjects in enumerate(ws['A1':'G6']):
    for n, cellObj in enumerate(rowOfCellObjects):
        cellObj.font = Font(size=20)

col_width = []
for i in range(len(next(ws.iter_rows()))):
    col_letter = get_column_letter(i + 1)

    minimum_width = 100
    current_width = ws.column_dimensions[col_letter].width
    if not current_width or current_width < minimum_width:
        ws.column_dimensions[col_letter].width = minimum_width

    col_width.append(ws.column_dimensions[col_letter].width)

for i, row in enumerate(ws):
    default_height = 26

    multiples_of_font_size = [default_height]
    for j, cell in enumerate(row):
        vertical = "center"
        horizontal = "center"
        if cell.value is not None:
            mul = 0
            for v in str(cell.value).split('\n'):
                mul += math.ceil(len(v) / col_width[j]) * cell.font.size

            if mul > 0:
                multiples_of_font_size.append(mul)

        cell.alignment = Alignment(vertical=vertical, horizontal=horizontal)

    original_height = ws.row_dimensions[i + 1].height
    if original_height is None:
        original_height = default_height

    new_height = max(multiples_of_font_size)
    if original_height < new_height:
        ws.row_dimensions[i + 1].height = new_height

wb.save("ОтчётКабинеты.xlsx")
wb.close()